"""Enumerated field counts per facility, read from the uploaded enumeration workbook.

This replaces the point-level synthetic register. The sheet is an aggregate: it reports
how many households were visited and how many eligible children were found per facility,
with no coordinates. That difference drives the whole comparison downstream - counts join
onto catchments by facility name, not by spatial containment.

Units matter here. ``registered_households`` counts enumeration *records* (households),
``registered_under5`` counts children 0-59 months. There is no headcount of people in the
sheet, so ``registered_population`` is derived from households and flagged as such; only
the household and under-5 figures are directly measured.
"""

import re
import unicodedata
from dataclasses import dataclass

import pandas as pd

import config
from sources import facilities

# Columns the sheet did not explain are kept under this prefix so they can be written
# back out untouched next to the engine's own columns.
PASSTHROUGH_PREFIX = "src::"

COUNT_COLUMNS = [
    "census_records", "eligible_children", "invited_members", "absent_households",
    "users_total", "users_active",
]
DENOMINATOR_COLUMNS = ["official_population", "official_target"]

OUTPUT_COLUMNS = [
    "facility_name", "registered_households", "registered_under5", "registered_population",
    "population_is_derived", "official_population", "official_target",
    "users_total", "users_active", "active_user_pct",
    "invited_members", "absent_households", "is_pooled", "pooled_with",
]


def has_enumeration(path):
    """True when an enumeration workbook was supplied for this run."""
    return path is not None and str(path).strip() != ""


def _fold(text):
    """Accent-free uppercase form with punctuation collapsed, for header matching."""
    text = unicodedata.normalize("NFKD", str(text))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()


def _match_headers(row):
    """Map ``{our column: position}`` for every known alias found in a candidate header row."""
    found = {}
    for position, cell in enumerate(row):
        if pd.isna(cell):
            continue
        folded = _fold(cell)
        if not folded:
            continue
        for column, aliases in config.ENUMERATION_COLUMNS.items():
            if column not in found and any(alias in folded for alias in aliases):
                found[column] = position
                break
    return found


def _find_header(raw):
    """Best header row in the first few rows: the one resolving the most known columns.

    A workbook commonly opens with a title row ("Until June 4 ... at 18:40"), so the real
    header is not row 0 and cannot be assumed to sit at any fixed offset.
    """
    best, best_row = {}, None
    for index in range(min(config.ENUMERATION_HEADER_SCAN_ROWS, len(raw))):
        found = _match_headers(raw.iloc[index])
        if len(found) > len(best):
            best, best_row = found, index
    return best_row, best


def _roster_match_score(values, roster_keys):
    """Fraction of a column's non-empty values that name an area in the roster."""
    candidates = [facilities.normalize(value) for value in values if pd.notna(value)]
    candidates = [key for key in candidates if key]
    if not candidates:
        return 0.0
    hits = sum(1 for key in candidates if key in roster_keys)
    return hits / len(candidates)


def _detect_name_column(body, roster):
    """The column naming each area, found by matching values against the geojson roster.

    Header text is unreliable across deployments - it may be in any language, or absent.
    The area names themselves are not: the upload already tells us what they are, so the
    column that contains them is identifiable whatever its header says.
    """
    if not roster:
        return None, 0.0
    roster_keys = {facilities.normalize(name) for name in roster}
    roster_keys.discard("")

    best_position, best_score = None, 0.0
    for position in range(body.shape[1]):
        score = _roster_match_score(body.iloc[:, position], roster_keys)
        if score > best_score:
            best_position, best_score = position, score
    if best_score < config.ENUMERATION_MIN_NAME_MATCH:
        return None, best_score
    return best_position, best_score


def _read_sheet(path, sheet_name, roster=None):
    """``(frame, columns)`` for one sheet, or ``(None, None)`` when it holds no area rows.

    Every column is carried through, not only the recognised ones: a spreadsheet from an
    unfamiliar deployment still has to join to its areas and come back with the engine's
    columns appended, even when nothing about its own metrics is understood.
    """
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
    if raw.empty:
        return None, None

    header_row, columns = _find_header(raw)
    header_row = 0 if header_row is None else header_row
    body = raw.iloc[header_row + 1:].reset_index(drop=True)

    header = [_fold(cell) for cell in raw.iloc[header_row]]
    if any(alias in cell for cell in header for alias in config.ENUMERATION_PER_USER_HEADERS):
        return None, None

    # Prefer the column that actually contains the area names over the one whose header
    # merely looks like a name column - the values are evidence, the header is a guess.
    detected, score = _detect_name_column(body, roster)
    if detected is not None:
        columns = {**columns, "facility_name": detected}
    elif "facility_name" not in columns:
        return None, None

    picked = pd.DataFrame({
        column: body.iloc[:, position] for column, position in columns.items()
    })
    # Original headers for every other column, so unrecognised metrics survive the trip.
    for position, label in enumerate(raw.iloc[header_row]):
        if position in set(columns.values()) or pd.isna(label):
            continue
        picked[f"{PASSTHROUGH_PREFIX}{label}"] = body.iloc[:, position]

    picked = picked.dropna(subset=["facility_name"])
    labels = picked["facility_name"].map(_fold)
    picked = picked[~labels.isin(config.ENUMERATION_TOTAL_LABELS)]
    for column in COUNT_COLUMNS + DENOMINATOR_COLUMNS:
        if column in picked.columns:
            picked[column] = pd.to_numeric(picked[column], errors="coerce")

    picked.attrs["name_match_score"] = score
    picked.attrs["layout"] = SheetLayout(
        sheet_name=sheet_name, header_row=header_row,
        name_position=columns["facility_name"], width=raw.shape[1])
    return picked.reset_index(drop=True), columns


@dataclass
class SheetLayout:
    """Where the area names sit in one sheet, so results can be written back beside them."""

    sheet_name: object
    header_row: int
    name_position: int
    width: int


def annotate_workbook(path, output_path, values_by_area, columns):
    """Copy the uploaded workbook to ``output_path`` with computed columns appended.

    The original file is returned intact - same sheets, same rows, same order, same
    formatting - with the engine's columns added to the right of each area's own row.
    That keeps the deliverable something the people who filled the sheet in already know
    how to read, rather than a separate export they have to reconcile against it.

    ``values_by_area`` maps a normalised area key to ``{column: value}``; ``columns`` fixes
    the order the new columns are written in.
    """
    from openpyxl import load_workbook

    # data_only=True substitutes each formula with the value the spreadsheet last computed.
    # Without it openpyxl keeps the formula but discards its cached result, so every total
    # in the file reads as blank to anything that does not recalculate - the uploaded
    # workbook would come back looking emptier than it went in.
    workbook = load_workbook(path, data_only=True)
    annotated = 0
    for layout in _layouts(path, values_by_area):
        sheet = workbook[layout.sheet_name] if layout.sheet_name in workbook.sheetnames \
            else workbook.worksheets[0]
        first_column = sheet.max_column + 1
        for offset, column in enumerate(columns):
            sheet.cell(row=layout.header_row + 1, column=first_column + offset, value=column)

        for row in range(layout.header_row + 2, sheet.max_row + 1):
            name = sheet.cell(row=row, column=layout.name_position + 1).value
            values = values_by_area.get(facilities.normalize(name))
            if not values:
                continue
            for offset, column in enumerate(columns):
                sheet.cell(row=row, column=first_column + offset, value=_cell(values.get(column)))
            annotated += 1

    workbook.save(output_path)
    return output_path, annotated


def _cell(value):
    """Excel-safe scalar: numpy and pandas types are not writable by openpyxl."""
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return None
    if isinstance(value, (bool, str)):
        return value
    if hasattr(value, "item"):
        return value.item()
    return value


def _layouts(path, values_by_area):
    """Layout of every sheet in the workbook that holds area rows."""
    roster = list(values_by_area)
    sheets = pd.read_excel(path, sheet_name=None, header=None)
    layouts = []
    for name in sheets:
        frame, _ = _read_sheet(path, name, roster)
        if frame is not None and "layout" in frame.attrs:
            layouts.append(frame.attrs["layout"])
    return layouts


def _merge_sheets(frames):
    """Fold every facility sheet into one row per facility.

    A workbook commonly holds several snapshots of the same campaign (a 15:00 sheet and an
    18:40 sheet in the reference file). Counts take the largest value seen, since a later
    snapshot supersedes an earlier one and enumeration only accumulates; official
    denominators take the first value seen, since they come from the microplan and do not
    move during a campaign.
    """
    combined = pd.concat(frames, ignore_index=True)
    combined["_key"] = combined["facility_name"].map(facilities.normalize)
    combined = combined[combined["_key"] != ""]

    aggregation = {"facility_name": "first"}
    for column in COUNT_COLUMNS:
        if column in combined.columns:
            aggregation[column] = "max"
    for column in DENOMINATOR_COLUMNS:
        if column in combined.columns:
            aggregation[column] = "first"
    # Unrecognised columns are passed straight through; the first value seen wins, since
    # nothing is known about whether they accumulate.
    for column in combined.columns:
        if column.startswith(PASSTHROUGH_PREFIX):
            aggregation[column] = "first"

    merged = combined.groupby("_key", as_index=False, sort=False).agg(aggregation)
    return merged.drop(columns="_key")


def load_enumeration(path, roster=None, pooled_policy=None, avg_household_size=None):
    """``(frame, resolution)``: enumerated counts per facility, resolved onto ``roster``.

    ``roster`` is the facility name list from the uploaded boundary geojson; when given,
    each row's ``facility_name`` is rewritten to the roster's spelling so the join is exact,
    and the returned :class:`~sources.facilities.Resolution` reports names that did not
    match in either direction. Rows pooling several facilities are handled per
    ``pooled_policy`` (default ``config.ENUMERATION_POOLED_POLICY``).

    ``avg_household_size`` must be the same size the estimate was computed at. It only
    affects the derived ``registered_population``; passing a different one would put the
    two sides of every population coverage ratio in different units.
    """
    sheets = pd.read_excel(path, sheet_name=None, header=None)
    frames = [frame for frame, _ in (_read_sheet(path, name, roster) for name in sheets)
              if frame is not None and not frame.empty]
    if not frames:
        raise ValueError(
            f"{path}: no sheet in this workbook has a column of area names. Expected either "
            f"a recognised header ({', '.join(config.ENUMERATION_COLUMNS['facility_name'])}) "
            f"or a column whose values match the uploaded geojson's areas.")

    source = _merge_sheets(frames)
    for column in COUNT_COLUMNS + DENOMINATOR_COLUMNS:
        if column not in source.columns:
            source[column] = pd.NA

    if roster is None:
        rows = source.copy()
        rows["is_pooled"] = False
        rows["pooled_with"] = None
        return _finalize(rows, avg_household_size), facilities.Resolution()

    resolution = facilities.resolve(source["facility_name"], roster)
    rows = source[source["facility_name"].isin(resolution.mapping)].copy()
    rows["facility_name"] = rows["facility_name"].map(resolution.mapping)
    rows["is_pooled"] = False
    rows["pooled_with"] = None
    policy = pooled_policy or config.ENUMERATION_POOLED_POLICY
    _apply_pooled(rows, source, resolution, policy)

    frame = _finalize(rows, avg_household_size)
    frame.attrs["withheld"] = _withheld(source, resolution, policy)
    return frame, resolution


def _withheld(source, resolution, policy):
    """Counts left out of the per-facility totals because a row pooled several facilities.

    Under the ``flag`` policy these records belong to no single facility and are excluded
    rather than invented onto one. They are reported so the shortfall against the sheet's
    own TOTAL row is explained rather than mysterious.
    """
    if policy == "apportion" or not resolution.pooled:
        return {}
    pooled = source[source["facility_name"].isin(resolution.pooled)]
    if pooled.empty:
        return {}
    return {
        "rows": list(resolution.pooled),
        "facilities": sorted({name for names in resolution.pooled.values() for name in names}),
        "registered_households": int(pooled["census_records"].fillna(0).sum()),
        "registered_under5": int(pooled["eligible_children"].fillna(0).sum()),
    }


def _target_shares(rows, members):
    """Each member's share of the pooled counts, by official target (equal split if absent)."""
    shares = {}
    for member in members:
        match = rows.loc[rows["facility_name"] == member, "official_target"]
        target = match.iloc[0] if not match.empty else None
        shares[member] = 0.0 if target is None or pd.isna(target) else float(target)
    total = sum(shares.values())
    if total <= 0:
        return {member: 1.0 / len(members) for member in members}
    return {member: share / total for member, share in shares.items()}


def _apply_pooled(rows, source, resolution, policy):
    """Attach pooled-row counts to their member facilities, in place, per ``policy``."""
    for pooled_name, members in resolution.pooled.items():
        pooled = source[source["facility_name"] == pooled_name]
        if pooled.empty:
            continue
        pooled = pooled.iloc[0]
        rows.loc[rows["facility_name"].isin(members), ["is_pooled", "pooled_with"]] = [
            True, ", ".join(members)]

        # Under the default "flag" policy the counts stay on neither member, so a pooled
        # facility reads as "not separable" rather than as a fabricated split.
        if policy != "apportion":
            continue

        shares = _target_shares(rows, members)
        for member, share in shares.items():
            mask = rows["facility_name"] == member
            for column in COUNT_COLUMNS:
                value = pooled.get(column)
                if not pd.isna(value):
                    rows.loc[mask, column] = round(float(value) * share)


def _finalize(rows, avg_household_size=None):
    """Derive the reported measures and return the stable output schema."""
    frame = rows.copy()
    households = frame["census_records"].fillna(0)
    household_size = avg_household_size or config.AVG_HOUSEHOLD_SIZE

    frame["registered_households"] = households.astype(int)
    frame["registered_under5"] = frame["eligible_children"].fillna(0).astype(int)
    # No headcount of people exists in the sheet; this is households scaled by the same
    # average size the estimate used, and is labelled derived wherever it is displayed.
    frame["registered_population"] = (households * household_size).round().astype(int)
    frame["population_is_derived"] = True

    active = frame["users_active"]
    total = frame["users_total"]
    frame["active_user_pct"] = (active / total).where(total > 0).round(4)

    for column in OUTPUT_COLUMNS:
        if column not in frame.columns:
            frame[column] = pd.NA
    passthrough = [c for c in frame.columns if c.startswith(PASSTHROUGH_PREFIX)]
    return frame[OUTPUT_COLUMNS + passthrough].reset_index(drop=True)


def main():
    import argparse

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="uploaded enumeration .xlsx")
    parser.add_argument("--boundaries", help="boundary geojson supplying the facility roster")
    args = parser.parse_args()

    roster = None
    if args.boundaries:
        from sources.catchments import load_boundary_upload

        roster = list(load_boundary_upload(args.boundaries).units["facility_name"])

    frame, resolution = load_enumeration(args.workbook, roster)
    print(resolution.summary())
    if resolution.unmatched_source:
        print("\nsheet rows with no matching facility:")
        for name in resolution.unmatched_source:
            hint = resolution.suggestions.get(name)
            print(f"  {name}" + (f"   (did you mean {hint}?)" if hint else ""))
    if resolution.unmatched_roster:
        print("\nfacilities with no enumeration row:")
        for name in resolution.unmatched_roster:
            print(f"  {name}")

    print(f"\nfacilities loaded:        {len(frame):>12,}")
    print(f"households enumerated:    {frame['registered_households'].sum():>12,}")
    print(f"children 0-59mo found:    {frame['registered_under5'].sum():>12,}")
    if frame["official_target"].notna().any():
        print(f"official target (cible):  {int(frame['official_target'].sum()):>12,}")
    withheld = frame.attrs.get("withheld") or {}
    if withheld:
        print(f"\npooled rows, counts withheld from the per-facility totals above:")
        print(f"  rows:       {', '.join(withheld['rows'])}")
        print(f"  facilities: {', '.join(withheld['facilities'])}")
        print(f"  withheld:   {withheld['registered_households']:,} households, "
              f"{withheld['registered_under5']:,} children")


if __name__ == "__main__":
    main()
