"""
Generate Chad boundary Excel in the same format as NIGERIA.xlsx.

Columns: Country, Province, District, Arrondissement, Health Facility, Village,
         Service Boundary Code, Boundary (French), Boundary (Arabic), Latitude, Longitude

Uses the 45 boundary codes from the synthetic dataset with N'Djamena hierarchy.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import csv
import numpy as np

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(SCRIPT_DIR, "individuals_flat.csv")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "..", "boundary_data", "CHAD.xlsx")

# N'Djamena arrondissement mapping (from boundary code structure)
# POLIO_CHAD_CH_01_XX where XX is the arrondissement number
ARRONDISSEMENTS = {
    "01": "1er Arrondissement",
    "02": "2ème Arrondissement",
    "03": "3ème Arrondissement",
    "04": "4ème Arrondissement",
    "05": "5ème Arrondissement",
    "06": "6ème Arrondissement",
    "07": "7ème Arrondissement",
    "08": "8ème Arrondissement",
    "10": "10ème Arrondissement",
}

# French/Arabic names for settlements
FRENCH_NAMES = {
    "AMTOUKOUI": "Amtoukoui",
    "SABANGALI": "Sabangali",
    "BOLOLO": "Bololo",
    "PARIS CONGO": "Paris Congo",
    "MARDJANDAFACK": "Mardjandafack",
    "KABALAYE": "Kabalayé",
    "GARDOLE": "Gardolé",
    "AMRIGUEBE": "Amriguébé",
    "ARDEP-TIMANE": "Ardep-Timaně",
    "MOURSAL": "Moursal",
    "REPOS": "Repos",
    "KLEMAT": "Klémat",
    "DEMBE": "Dembé",
    "RIDINA": "Ridina",
    "CHAGOUA": "Chagoua",
    "NDJAMENA CENTRE": "N'Djaména Centre",
    "ABENA": "Abéna",
    "NDJARI": "N'Djari",
    "HABENA": "Habéna",
    "ANGABO": "Angabo",
    "DIGUEL": "Diguel",
    "AMBASSATNA": "Ambassatna",
    "GASSI": "Gassi",
    "ALBIR": "Albir",
    "AL-AFIA": "Al-Afia",
    "AMBATTA 2": "Ambatta 2",
    "AMBATTA 1": "Ambatta 1",
    "ATRONE": "Atroné",
    "KOUNDOUL": "Koundoul",
    "NDJAMENA FARA": "N'Djaména Fara",
    "NGOUNBA MASSA": "Ngounba Massa",
    "NGUELI": "Nguéli",
    "ORDRE DE MALTE": "Ordre de Malte",
    "SAINTE THERESE": "Sainte Thérèse",
    "TOUKRA I": "Toukra I",
    "TOUKRA II": "Toukra II",
    "TOUKRA III": "Toukra III",
    "WALIA EST": "Walia Est",
    "CARRE 5 POL 50": "Carré 5 Pol 50",
    "CARRE 6 POL 46": "Carré 6 Pol 46",
    "CARRE 6 POL 47": "Carré 6 Pol 47",
    "CARRE 7 POL 48": "Carré 7 Pol 48",
    "CARRE 8 POL 51": "Carré 8 Pol 51",
    "CARRE 8 POL 52": "Carré 8 Pol 52",
    "CARRE 9 POL 42": "Carré 9 Pol 42",
}

ARABIC_NAMES = {
    "AMTOUKOUI": "أمتوكوي",
    "SABANGALI": "سابنقالي",
    "BOLOLO": "بولولو",
    "PARIS CONGO": "باريس كونغو",
    "MARDJANDAFACK": "مردجندافاك",
    "KABALAYE": "كبلاي",
    "GARDOLE": "قاردولي",
    "AMRIGUEBE": "أمريقيبي",
    "ARDEP-TIMANE": "أرديب تيمان",
    "MOURSAL": "مورسال",
    "REPOS": "ريبو",
    "KLEMAT": "كليمات",
    "DEMBE": "ديمبي",
    "RIDINA": "ريدينا",
    "CHAGOUA": "شاقوا",
    "NDJAMENA CENTRE": "نجامينا المركز",
    "ABENA": "أبينا",
    "NDJARI": "نجاري",
    "HABENA": "حبينا",
    "ANGABO": "أنقابو",
    "DIGUEL": "ديقيل",
    "AMBASSATNA": "أمبستنا",
    "GASSI": "قاسي",
    "ALBIR": "البير",
    "AL-AFIA": "العافية",
    "AMBATTA 2": "أمبطة 2",
    "AMBATTA 1": "أمبطة 1",
    "ATRONE": "أتروني",
    "KOUNDOUL": "كوندول",
    "NDJAMENA FARA": "نجامينا فارا",
    "NGOUNBA MASSA": "نقونبا ماسا",
    "NGUELI": "نقيلي",
    "ORDRE DE MALTE": "أمر مالطا",
    "SAINTE THERESE": "سانت تيريز",
    "TOUKRA I": "توكرا 1",
    "TOUKRA II": "توكرا 2",
    "TOUKRA III": "توكرا 3",
    "WALIA EST": "واليا الشرقية",
    "CARRE 5 POL 50": "كاريه 5 بول 50",
    "CARRE 6 POL 46": "كاريه 6 بول 46",
    "CARRE 6 POL 47": "كاريه 6 بول 47",
    "CARRE 7 POL 48": "كاريه 7 بول 48",
    "CARRE 8 POL 51": "كاريه 8 بول 51",
    "CARRE 8 POL 52": "كاريه 8 بول 52",
    "CARRE 9 POL 42": "كاريه 9 بول 42",
}


def get_boundary_centroids():
    """Get average lat/lon per boundary from CSV."""
    centroids = {}
    counts = {}
    with open(CSV_PATH, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = row["boundary_code"]
            lat = float(row["latitude"])
            lon = float(row["longitude"])
            name = row["locality_name"]
            if code not in centroids:
                centroids[code] = {"lat": 0, "lon": 0, "name": name}
                counts[code] = 0
            centroids[code]["lat"] += lat
            centroids[code]["lon"] += lon
            counts[code] += 1

    for code in centroids:
        centroids[code]["lat"] = round(centroids[code]["lat"] / counts[code], 6)
        centroids[code]["lon"] = round(centroids[code]["lon"] / counts[code], 6)

    return centroids


def parse_boundary_code(code):
    """Extract hierarchy from boundary code."""
    # POLIO_CHAD_CH_01_XX_YY_CS_NAME or POLIO_CHAD_CH_01_10_17_09_CARR__6_POL_46
    parts = code.replace("POLIO_CHAD_CH_01_", "").split("_")

    arrond_num = parts[0]  # e.g., "01", "10"

    if "CS" in parts:
        # Health center: POLIO_CHAD_CH_01_XX_YY_CS_NAME
        facility_idx = parts.index("CS")
        hf_num = parts[facility_idx - 1]
        name_parts = parts[facility_idx + 1:]
        name = " ".join(name_parts).replace("  ", " ")
        return arrond_num, hf_num, name, "CS"
    elif "CARR" in parts:
        # Polling area
        name_parts = [p for p in parts[2:] if p]
        name = " ".join(name_parts).replace("  ", " ").replace("CARR ", "CARRE ").replace("POL ", "POL ")
        return arrond_num, parts[1], name, "CARR"
    else:
        return arrond_num, parts[1] if len(parts) > 1 else "01", " ".join(parts[1:]), "OTHER"


def generate_excel():
    centroids = get_boundary_centroids()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Boundary Data"

    # Headers
    headers = [
        "Country", "Province", "District", "Arrondissement",
        "Health Facility", "Village/Settlement",
        "Service Boundary Code",
        "Boundary (French)", "Boundary (Arabic)",
        "Latitude", "Longitude"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 55
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 12

    row_num = 2

    def write_row(values):
        nonlocal row_num
        for col, val in enumerate(values, 1):
            ws.cell(row=row_num, column=col, value=val)
        row_num += 1

    # Country level
    write_row(["Chad", "", "", "", "", "", "POLIO_CHAD_CH",
               "Tchad", "تشاد", None, None])

    # Province level
    write_row(["Chad", "N'Djaména", "", "", "", "", "POLIO_CHAD_CH_01",
               "N'Djaména", "نجامينا", None, None])

    # District level
    write_row(["Chad", "N'Djaména", "N'Djaména Ville", "", "", "",
               "POLIO_CHAD_CH_01", "N'Djaména Ville", "مدينة نجامينا",
               None, None])

    # Group boundaries by arrondissement
    arrond_groups = {}
    for code, info in centroids.items():
        arrond_num, hf_num, name, btype = parse_boundary_code(code)
        if arrond_num not in arrond_groups:
            arrond_groups[arrond_num] = []
        arrond_groups[arrond_num].append((code, info, hf_num, name, btype))

    # Sort and write
    for arrond_num in sorted(arrond_groups.keys()):
        arrond_name = ARRONDISSEMENTS.get(arrond_num, f"{arrond_num}ème Arrondissement")
        arrond_code = f"POLIO_CHAD_CH_01_{arrond_num}"

        # Arrondissement row
        write_row(["Chad", "N'Djaména", "N'Djaména Ville", arrond_name, "", "",
                   arrond_code, arrond_name, "", None, None])

        # Health facilities / settlements under this arrondissement
        boundaries = sorted(arrond_groups[arrond_num], key=lambda x: x[0])
        for code, info, hf_num, name, btype in boundaries:
            locality = info["name"]
            french = FRENCH_NAMES.get(locality, locality.title())
            arabic = ARABIC_NAMES.get(locality, "")

            if btype == "CS":
                facility_name = f"CS {french}"
                write_row(["Chad", "N'Djaména", "N'Djaména Ville", arrond_name,
                           facility_name, "", code, french, arabic, None, None])
                # Village row under the facility
                write_row(["Chad", "N'Djaména", "N'Djaména Ville", arrond_name,
                           facility_name, locality, code, french, arabic,
                           info["lat"], info["lon"]])
            else:
                # Polling area - treat as village directly under arrondissement
                write_row(["Chad", "N'Djaména", "N'Djaména Ville", arrond_name,
                           "", locality, code, locality.title(), arabic,
                           info["lat"], info["lon"]])

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Excel saved: {OUTPUT_PATH}")
    print(f"Total rows: {row_num - 1} (including header)")


if __name__ == "__main__":
    generate_excel()
